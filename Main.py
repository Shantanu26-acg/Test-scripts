from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from threading import Thread, Event, Lock
from SubjectCreation import subject_creation
from SubjectActivation import activate_subject
from LocationCreation import create_location
from ProductCreation import create_and_activate_product
from LineCreation import line_creation
from POCreation import create_po
from openpyxl import load_workbook
import time
import pandas as pd
import queue
import threading

chrome_options=Options()
chrome_options.add_argument('--ignore-certificate-errors')
chrome_options.add_argument('ignore-ssl-errors')
chrome_options.add_argument('allow-insecure-localhost')
chrome_options.add_argument('--disable-web-security')
chrome_options.add_argument('--disable-blink-features=AutomationControlled')

service=Service(r"C:\Users\Shantanu\Downloads\chromedriver-win32\chromedriver.exe")

driver_creation=webdriver.Chrome(service=service, options=chrome_options)
wait_creation=WebDriverWait(driver_creation, 15)

driver_activation=webdriver.Chrome(service=service, options=chrome_options)
wait_activation=WebDriverWait(driver_activation, 15)

driver_location_creation=webdriver.Chrome(service=service, options=chrome_options)
wait_location_creation=WebDriverWait(driver_location_creation, 15)

driver_product_creation=webdriver.Chrome(service=service, options=chrome_options)
wait_product_creation=WebDriverWait(driver_product_creation, 15)

driver_line_creation=webdriver.Chrome(service=service, options=chrome_options)
wait_line_creation=WebDriverWait(driver_line_creation, 15)

driver_po_creation=webdriver.Chrome(service=service, options=chrome_options)
wait_po_creation=WebDriverWait(driver_po_creation, 15)

df_urls=pd.read_excel(
    "Master data.xlsx",
    sheet_name="URLs",
    header=None
)

df_logindata=pd.read_excel(
    "Master data.xlsx",
    sheet_name="Login Data"
)

username_1=df_logindata.loc[0, "Usernames"]
password_1=df_logindata.loc[0, "Passwords"]

username_2=df_logindata.loc[1, "Usernames"]
password_2=df_logindata.loc[1, "Passwords"]

username_3=df_logindata.loc[2, 'Usernames']
password_3=df_logindata.loc[2, 'Passwords']

username_4=df_logindata.loc[3, 'Usernames']
password_4=df_logindata.loc[3, 'Passwords']

username_5=df_logindata.loc[4, 'Usernames']
password_5=df_logindata.loc[4, 'Passwords']

username_6=df_logindata.loc[5, 'Usernames']
password_6=df_logindata.loc[5, 'Passwords']

print(username_1, username_2, username_3, username_4, username_5, username_6, password_1, password_2, password_3, password_4, password_5, password_6)

login_url=df_urls.iloc[0, 0]
masterdata_url=df_urls.iloc[1, 0]

product_start_event=Event()
line_start_event=Event()
line_done_event=Event()
location_lock=Lock()
location_count=0


df=pd.read_excel("Master data.xlsx", sheet_name="Subject Master Data", engine="openpyxl", dtype={
    "GLN": str,
    "GS1 Company Prefix": str,
    "Remarks": str
})

# Ensure Remarks column exists and is of string type
if 'Remarks' not in df.columns:
    df['Remarks'] = ""
else:
    df['Remarks'] = df['Remarks'].astype(str)


dp=pd.read_excel(
    "Master data.xlsx",
    sheet_name="Product Master Data",
    dtype={
        "GS1 Company Prefix": str,
        "GTIN": str,
        "No. of items in this level": str,
        "SSCC Extenson Digit(Loose Pack)": str
    }
)

dp['Product No'] = dp['Product No'].ffill()


dl=pd.read_excel("Master data.xlsx", sheet_name="Line Data", dtype={
    "Location Number": str,
    "Line Number": str,
    "Systems": str
})

dl['Location Number'] = dl['Location Number'].ffill()
dl['Location Name'] = dl['Location Name'].ffill()
dl['Line Name'] = dl['Line Name'].ffill()
dl['Line number'] = dl['Line number'].ffill()

dc=pd.read_excel("Master Data.xlsx", sheet_name="PO Data", dtype={
    "Lot/Batch Number": str
})



batch_queue=queue.Queue()
location_queue=queue.Queue()

def write_remark_to_excel(subject_name, remark, file_path="Master data.xlsx", sheet_name="Subject Master Data"):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Get column indices for Subject Name and Remarks
    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    subject_col = headers.get("Subject Name")
    remarks_col = headers.get("Remarks")

    if not subject_col or not remarks_col:
        print("❌ 'Subject Name' or 'Remarks' column not found.")
        return

    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[subject_col - 1].value == subject_name:
            row[remarks_col - 1].value = remark
            break

    wb.save(file_path)
    wb.close()


def creation_thread():
    driver_creation.get(login_url)
    wait_creation.until(EC.visibility_of_element_located((By.NAME, "email"))).send_keys(username_1)
    driver_creation.find_element(By.NAME, "password").send_keys(password_1)
    driver_creation.find_element(By.XPATH, "//input[@type='submit' and @value='Login']").click()

    batch_names=[]
    for idx, row in df.iterrows():
        name=row['Subject Name']
        print(f"Creating subjects: {row['Subject Name']}")

        remarks = [""]

        subject_creation(driver_creation, wait_creation,
                         subject_name=name,
                         group=row['Subject Group'],
                         address=row['Address'],
                         city=row['City'],
                         zip=row['Zip'],
                         state=row['State'],
                         country=row['Country'],
                         gln=row['GLN'],
                         gs1=row['GS1 Company Prefix'],
                         remarks=remarks
        )

        write_remark_to_excel(name, remarks[0])
        batch_names.append(name)

        if len(batch_names)==3:
            print(f"Batch Ready for activation: {batch_names}")
            batch_queue.put(batch_names.copy())
            batch_names.clear()

    batch_names=[]

    if batch_names:
        print(f"Final batch ready for activation: {batch_names}")
        batch_queue.put(batch_names.copy())

    batch_queue.put(None)

    print(df[["Subject Name", "Remarks"]])

def activation_thread():
    driver_activation.get(login_url)
    wait_activation.until(EC.visibility_of_element_located((By.NAME, "email"))).send_keys(username_2)
    driver_activation.find_element(By.NAME, "password").send_keys(password_2)
    driver_activation.find_element(By.XPATH, "//input[@type='submit' and @value='Login']").click()

    while True:
        batch=batch_queue.get()
        if batch is None:
            break

        print(f"Activating subjects: {batch}")
        for name in batch:
            activate_subject(driver_activation, wait_activation, name=name)
            time.sleep(3)
            location_queue.put(name)

    print("All activations complete")

def location_thread():
    global location_count

    driver_location_creation.get(login_url)
    wait_location_creation.until(EC.visibility_of_element_located((By.NAME, "email"))).send_keys(username_3)
    driver_location_creation.find_element(By.NAME, "password").send_keys(password_3)
    driver_location_creation.find_element(By.XPATH, "//input[@type='submit' and @value='Login']").click()

    while True:
        name=location_queue.get()
        if name is None:
            break

        print(f"Creating location for: {name}")
        row_data=df[df['Subject Name']==name].iloc[0]

        create_location(
            driver=driver_location_creation,
            wait=wait_location_creation,
            location_name="loc_"+name,
            subject_name=name,
            address=row_data['Address'],
            city=row_data['City'],
            state=row_data['State'],
            zip=row_data['Zip'],
            country=row_data['Country'],
            gln=row_data['GLN']
        )

        with location_lock:
            location_count+=1
            if location_count==3:
                print("3 locations created, signaling product creation thread to start")
                product_start_event.set()
            if location_count == 9:
                print("9 locations created, signaling line creation thread to start")
                line_start_event.set()


def product_creation_thread():
    product_start_event.wait()

    driver_product_creation.get(login_url)
    wait_product_creation.until(EC.visibility_of_element_located((By.NAME, "email"))).send_keys(username_4)
    driver_product_creation.find_element(By.NAME, "password").send_keys(password_4)
    driver_product_creation.find_element(By.XPATH, "//input[@type='submit' and @value='Login']").click()

    create_and_activate_product(driver_product_creation, wait_product_creation, dp)

def line_thread():
    line_start_event.wait()

    driver_line_creation.get(login_url)
    wait_line_creation.until(EC.visibility_of_element_located((By.NAME, "email"))).send_keys(username_5)
    driver_line_creation.find_element(By.NAME, "password").send_keys(password_5)
    driver_line_creation.find_element(By.XPATH, "//input[@type='submit' and @value='Login']").click()

    grouped = dl.groupby(['Location Number', 'Location Name', 'Line Name', 'Line number'])

    for (loc_no, loc_name, line_name, line_key), group_df in grouped:
        systems = group_df['System Type'].dropna().tolist()
        no_of_systems = len(systems)

        print(f"⚙️ Creating line '{line_name}' for location '{loc_name}' with systems: {systems}")

        try:
            line_creation(
                driver=driver_line_creation,
                wait=wait_line_creation,
                loc_name=loc_name,
                line_name=line_name,
                line_key=line_key,
                no_of_systems=no_of_systems,
                systems=systems
            )
        except Exception as e:
            print(f"❌ Failed to create line for {loc_name} - {e}")

    line_done_event.set()
    print("✅ Line creation complete.")

def po_thread():
    line_done_event.wait()

    driver_po_creation.get(login_url)
    wait_po_creation.until(EC.visibility_of_element_located((By.NAME, "email"))).send_keys(username_6)
    driver_po_creation.find_element(By.NAME, "password").send_keys(password_6)
    driver_po_creation.find_element(By.XPATH, "//input[@type='submit' and @value='Login']").click()

    print("Available columns in Excel:", dc.columns.tolist())

    for index, row in dc.iterrows():
        print(f"\nProcessing PO #{index + 1} — {row['PO Name']}")
        for col in dc.columns:
            print(repr(col))

        create_po(
            driver_po_creation,
            wait_po_creation,
            po_name=row['PO Name'],
            serial_numbers_provider=row['Serial Numbers Provider'],
            product_name=row['Product'],
            packaging_version=row['Packaging Version'],
            lot_number=row['Lot/Batch Number'],
            production_date=row['Production Date'],
            mfg_date=row['MFG Date'],
            expiration_date=row['Expiration Date'],
            regulation=row['Regulation'],
            manufacturing_location=row['Manufacturing Location'],
            quantity=row['Quantity'],
            production_line=str(row['Production Line'])
        )

    time.sleep(5)


t1=threading.Thread(target=creation_thread)
t2=threading.Thread(target=activation_thread)
t3=threading.Thread(target=location_thread)
t4=threading.Thread(target=product_creation_thread)
t5=threading.Thread(target=line_thread)
t6=threading.Thread(target=po_thread)

t1.start()
t2.start()
t3.start()
t4.start()
t5.start()
t6.start()

t1.join()
t2.join()
t3.join()
t4.join()
t5.join()
t6.join()

driver_creation.quit()
driver_activation.quit()
driver_location_creation.quit()
driver_product_creation.quit()
driver_line_creation.quit()
driver_po_creation.quit()

time.sleep(5)